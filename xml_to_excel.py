import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Ask user if they want to include logs for all tests or only failed ones
user_input = input("Do you want to include logs for all test cases? (y/N): ").strip().lower()
include_all_logs = user_input == "y"

tree = ET.parse("results.xml")
root = tree.getroot()
namespaces = {"": "http://www.froglogic.com/resources/schemas/xml3"}


def convert_to_datetime(time_str):
    return datetime.strptime(time_str.replace("+05:30", "").replace("Z", ""), "%Y-%m-%dT%H:%M:%S.%f")


results_by_module = {}
serial_number = 1

for testcase in root.findall(".//test[@type='testcase']", namespaces):
    module_name = None
    for log in testcase.findall(".//message", namespaces):
        if log.find("text", namespaces) is not None:
            msg_text = log.find("text", namespaces).text.strip()
            if msg_text.lower() in ["login", "recipe"]:
                module_name = msg_text
                break
    module_name = module_name or "Unknown"

    for section in testcase.findall(".//test[@type='section']", namespaces):
        test_name = section.find("prolog/name", namespaces).text
        start_time = section.find("prolog", namespaces).attrib["time"]
        end_time = section.find("epilog", namespaces).attrib["time"]
        start_dt = convert_to_datetime(start_time)
        end_dt = convert_to_datetime(end_time)

        logs = []
        for message in section.findall(".//message", namespaces):
            if message.find("text", namespaces) is not None:
                logs.append(message.find("text", namespaces).text.strip())

        verification = section.find(".//verification/scriptedVerificationResult", namespaces)
        status = verification.attrib["type"] if verification is not None else "FAIL"

        result = {
            "Sr No.": serial_number,
            "Test Case Name": test_name,
            "Start Time": start_time,
            "End Time": end_time,
            "Total Time (seconds)": (end_dt - start_dt).total_seconds(),
            "Status": status,
            "Logs": "\n".join(logs) if include_all_logs or status == "FAIL" else "",
            "Comments": "",
        }
        results_by_module.setdefault(module_name, []).append(result)
        serial_number += 1

# Create workbook
wb = Workbook()
del wb["Sheet"]

# Define styles
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
pass_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
bold_font_white = Font(bold=True, color="FFFFFF")
bold_font = Font(bold=True)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

summary_data = []

# Create individual module sheets
for module, records in results_by_module.items():
    df = pd.DataFrame(records)
    pass_count = (df["Status"] == "PASS").sum()
    fail_count = (df["Status"] == "FAIL").sum()
    summary_data.append({"Module Name": module, "Total Test Cases": len(df), "Pass": pass_count, "Fail": fail_count})

    ws = wb.create_sheet(module)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font_white
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):  # Status Column
        for cell in row:
            if cell.value == "PASS":
                cell.fill = pass_fill
                cell.font = bold_font
            elif cell.value == "FAIL":
                cell.fill = fail_fill
                cell.font = bold_font

# Create Summary Sheet
summary_df = pd.DataFrame(summary_data)
ws_summary = wb.create_sheet("Summary")
for r in dataframe_to_rows(summary_df, index=False, header=True):
    ws_summary.append(r)

# Style summary headers
for cell in ws_summary[1]:
    cell.fill = header_fill
    cell.font = bold_font_white
    cell.alignment = Alignment(horizontal="center")

# Color specific columns
ws_summary["C1"].fill = pass_fill  # Pass = Green
ws_summary["D1"].fill = fail_fill  # Fail = Red

# Center-align and border summary data
for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

# Add total row
ws_summary.append([""])
total_row = ["Total", sum(x["Total Test Cases"] for x in summary_data), sum(x["Pass"] for x in summary_data), sum(x["Fail"] for x in summary_data)]
ws_summary.append(total_row)

# Bold total row
for cell in ws_summary[ws_summary.max_row]:
    cell.font = Font(bold=True)

# Add chart
chart = BarChart()
chart.title = "Test Results Summary"
chart.x_axis.title = "Module"
chart.y_axis.title = "Number of Test Cases"
chart.style = 10
chart.type = "col"

data_ref = Reference(ws_summary, min_col=3, min_row=1, max_col=4, max_row=ws_summary.max_row - 2)
cat_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=ws_summary.max_row - 2)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)

for series in chart.series:
    series.dLbls = DataLabelList()
    series.dLbls.show_val = True

ws_summary.add_chart(chart, "F2")

# Save the file
wb.save("test_report_with_modules_summary.xlsx")
print("✅ Excel report 'test_report_with_modules_summary.xlsx' generated successfully!")
